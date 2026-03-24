"""
Export router for generating PDF, Excel, and CSV reports.
"""

import io
import csv
from datetime import datetime
from typing import Annotated, Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from database import get_db
from models import User, Competition, WOD, Athlete
from auth import get_current_judge_or_admin
from scoring import get_competition_leaderboard, format_result

router = APIRouter(prefix="/api/export", tags=["Export"])


@router.get("/{competition_id}/excel")
async def export_excel(
    competition_id: int,
    current_user: Annotated[User, Depends(get_current_judge_or_admin)],
    db: AsyncSession = Depends(get_db),
    divisions: Optional[str] = Query(None, description="Comma-separated divisions"),
):
    """
    Export competition results to Excel.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl"
        )

    # Get competition
    comp_result = await db.execute(
        select(Competition).where(Competition.id == competition_id)
    )
    competition = comp_result.scalar_one_or_none()
    if not competition:
        raise HTTPException(status_code=404, detail="Competition not found")

    # Get WODs
    wods_result = await db.execute(
        select(WOD)
        .where(WOD.competition_id == competition_id)
        .order_by(WOD.order_in_competition)
    )
    wods = wods_result.scalars().all()

    # Get leaderboards
    leaderboards = await get_competition_leaderboard(db, competition_id)

    # Filter divisions if specified
    if divisions:
        div_list = [d.strip() for d in divisions.split(",")]
        leaderboards = {k: v for k, v in leaderboards.items() if k in div_list}

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for division, entries in sorted(leaderboards.items()):
        # Create sheet for this division (max 31 chars for sheet name)
        sheet_name = division[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6 + len(wods))
        title_cell = ws.cell(row=1, column=1, value=f"{competition.name} - {division}")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # Header row
        headers = ["Pos", "Dorsal", "Nombre", "Box", "Total"]
        for wod in wods:
            headers.append(f"{wod.name}\nPts")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, entry in enumerate(entries, 4):
            ws.cell(row=row_idx, column=1, value=entry.rank).alignment = cell_alignment
            ws.cell(row=row_idx, column=2, value=entry.bib_number).alignment = cell_alignment
            ws.cell(row=row_idx, column=3, value=entry.athlete_name)
            ws.cell(row=row_idx, column=4, value=entry.box or "")
            ws.cell(row=row_idx, column=5, value=entry.total_points).alignment = cell_alignment

            # WOD scores
            wod_scores_map = {ws_data["wod_id"]: ws_data for ws_data in entry.wod_scores}
            for col_idx, wod in enumerate(wods, 6):
                ws_data = wod_scores_map.get(wod.id, {})
                points = ws_data.get("points", 0)
                cell = ws.cell(row=row_idx, column=col_idx, value=points)
                cell.alignment = cell_alignment

            # Add borders
            for col in range(1, 6 + len(wods)):
                ws.cell(row=row_idx, column=col).border = thin_border

        # Adjust column widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 10
        for i in range(len(wods)):
            ws.column_dimensions[chr(ord("F") + i)].width = 12

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{competition.name.replace(' ', '_')}_resultados_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{competition_id}/csv")
async def export_csv(
    competition_id: int,
    current_user: Annotated[User, Depends(get_current_judge_or_admin)],
    db: AsyncSession = Depends(get_db),
    division: Optional[str] = None,
):
    """
    Export competition results to CSV.
    """
    # Get competition
    comp_result = await db.execute(
        select(Competition).where(Competition.id == competition_id)
    )
    competition = comp_result.scalar_one_or_none()
    if not competition:
        raise HTTPException(status_code=404, detail="Competition not found")

    # Get WODs
    wods_result = await db.execute(
        select(WOD)
        .where(WOD.competition_id == competition_id)
        .order_by(WOD.order_in_competition)
    )
    wods = wods_result.scalars().all()

    # Get leaderboards
    leaderboards = await get_competition_leaderboard(db, competition_id)

    if division:
        if division not in leaderboards:
            raise HTTPException(status_code=404, detail="Division not found")
        leaderboards = {division: leaderboards[division]}

    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)

    # Headers
    headers = ["Division", "Posicion", "Dorsal", "Nombre", "Box", "Total"]
    for wod in wods:
        headers.extend([f"{wod.name}_Resultado", f"{wod.name}_Puntos"])
    writer.writerow(headers)

    # Data
    for div_name, entries in sorted(leaderboards.items()):
        for entry in entries:
            row = [
                div_name,
                entry.rank,
                entry.bib_number,
                entry.athlete_name,
                entry.box or "",
                entry.total_points,
            ]

            wod_scores_map = {ws_data["wod_id"]: ws_data for ws_data in entry.wod_scores}
            for wod in wods:
                ws_data = wod_scores_map.get(wod.id, {})
                result = ws_data.get("result")
                if result is not None:
                    formatted = format_result(result, wod.wod_type, ws_data.get("result_type", "RX"))
                else:
                    formatted = "-"
                row.extend([formatted, ws_data.get("points", 0)])

            writer.writerow(row)

    output.seek(0)

    filename = f"{competition.name.replace(' ', '_')}_resultados_{datetime.now().strftime('%Y%m%d')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{competition_id}/pdf")
async def export_pdf(
    competition_id: int,
    current_user: Annotated[User, Depends(get_current_judge_or_admin)],
    db: AsyncSession = Depends(get_db),
    division: Optional[str] = None,
):
    """
    Export competition results to PDF.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="reportlab not installed. Run: pip install reportlab"
        )

    # Get competition
    comp_result = await db.execute(
        select(Competition).where(Competition.id == competition_id)
    )
    competition = comp_result.scalar_one_or_none()
    if not competition:
        raise HTTPException(status_code=404, detail="Competition not found")

    # Get WODs
    wods_result = await db.execute(
        select(WOD)
        .where(WOD.competition_id == competition_id)
        .order_by(WOD.order_in_competition)
    )
    wods = wods_result.scalars().all()

    # Get leaderboards
    leaderboards = await get_competition_leaderboard(db, competition_id)

    if division:
        if division not in leaderboards:
            raise HTTPException(status_code=404, detail="Division not found")
        leaderboards = {division: leaderboards[division]}

    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        alignment=1,
        spaceAfter=20,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Heading2"],
        fontSize=12,
        alignment=1,
        spaceAfter=10,
    )

    elements = []

    for div_name, entries in sorted(leaderboards.items()):
        # Title
        elements.append(Paragraph(competition.name, title_style))
        elements.append(Paragraph(div_name, subtitle_style))

        # Table data
        headers = ["Pos", "Dorsal", "Nombre", "Box", "Total"]
        for wod in wods:
            headers.append(wod.name[:10])  # Truncate WOD names

        table_data = [headers]

        for entry in entries:
            row = [
                str(entry.rank),
                entry.bib_number,
                entry.athlete_name[:20],  # Truncate long names
                (entry.box or "")[:15],
                f"{entry.total_points:.0f}",
            ]

            wod_scores_map = {ws_data["wod_id"]: ws_data for ws_data in entry.wod_scores}
            for wod in wods:
                ws_data = wod_scores_map.get(wod.id, {})
                points = ws_data.get("points", 0)
                row.append(f"{points:.0f}")

            table_data.append(row)

        # Create table
        col_widths = [1.2*cm, 2*cm, 5*cm, 4*cm, 2*cm] + [2*cm] * len(wods)
        table = Table(table_data, colWidths=col_widths)

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 1*cm))

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    filename = f"{competition.name.replace(' ', '_')}_resultados_{datetime.now().strftime('%Y%m%d')}.pdf"

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{competition_id}/athletes/template")
async def get_athletes_import_template(
    competition_id: int,
    current_user: Annotated[User, Depends(get_current_judge_or_admin)],
    db: AsyncSession = Depends(get_db),
):
    """
    Download CSV template for athlete import.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Headers
    writer.writerow(["name", "gender", "birth_date", "division", "box", "bib_number"])

    # Example rows
    writer.writerow(["Juan Perez", "Masculino", "1990-05-15", "RX Masculino", "CrossFit Central", "001"])
    writer.writerow(["Maria Garcia", "Femenino", "1992-08-20", "RX Femenino", "CrossFit Norte", "002"])

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=plantilla_atletas.csv"},
    )
